from flask import Flask, render_template, request, jsonify, send_file
import pandas as pd
import requests
from concurrent.futures import ThreadPoolExecutor
import threading
import time
from openpyxl.styles import PatternFill
from io import BytesIO
import openpyxl
import os

app = Flask(__name__)

# SSL doğrulamasını kapatma (production'da açılmalı)
import urllib3
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

def check_url(url):
    try:
        start_time = time.time()
        
        # AllOrigins API kullanarak isteği yap
        proxy_url = f'https://api.allorigins.win/get?url={requests.utils.quote(url)}'
        
        headers = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36',
            'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8',
            'Accept-Language': 'tr-TR,tr;q=0.9,en-US;q=0.8,en;q=0.7'
        }
        
        response = requests.get(proxy_url, timeout=5, headers=headers)
        response_time = time.time() - start_time
        
        if response.status_code == 200:
            try:
                data = response.json()
                status_code = data.get('status', {}).get('http_code', 0)
                content = data.get('contents', '')
                
                # Çalışma durumunu belirle
                is_working = False
                
                # Çalışma kriterleri:
                # 1. AllOrigins yanıtı başarılı
                # 2. Orijinal site yanıtı başarılı
                # 3. İçerik var ve boş değil
                # 4. Yanıt süresi makul
                if (status_code == 200 and
                    content and
                    len(content.strip()) > 0 and
                    response_time < 5):
                    is_working = True
                
                return {
                    'url': url,
                    'status': f"{'Çalışıyor' if is_working else 'Çalışmıyor'} ({status_code})",
                    'working': is_working,
                    'response_time': f"{response_time:.2f} saniye",
                    'content_type': 'text/html',
                    'server': 'Remote Server',
                    'content': content[:200] if content else 'İçerik alınamadı'
                }
            except ValueError:
                return {
                    'url': url,
                    'status': 'Çalışmıyor (Yanıt Hatası)',
                    'working': False,
                    'response_time': f"{response_time:.2f} saniye",
                    'content_type': 'Bilinmiyor',
                    'server': 'Bilinmiyor',
                    'content': 'Yanıt işlenemedi'
                }
        else:
            return {
                'url': url,
                'status': f'Çalışmıyor ({response.status_code})',
                'working': False,
                'response_time': f"{response_time:.2f} saniye",
                'content_type': 'Bilinmiyor',
                'server': 'Bilinmiyor',
                'content': 'Proxy yanıt vermedi'
            }
            
    except requests.exceptions.Timeout:
        return {
            'url': url,
            'status': 'Çalışmıyor (Zaman Aşımı)',
            'working': False,
            'response_time': '> 5.00 saniye',
            'content_type': 'Bilinmiyor',
            'server': 'Bilinmiyor',
            'content': 'Sunucu yanıt vermedi (timeout)'
        }
    except (requests.exceptions.ConnectionError, requests.exceptions.RequestException) as e:
        return {
            'url': url,
            'status': 'Çalışmıyor (Bağlantı Hatası)',
            'working': False,
            'response_time': '-',
            'content_type': 'Bilinmiyor',
            'server': 'Bilinmiyor',
            'content': f'Sunucuya bağlanılamadı: {str(e)}'
        }

def process_urls(urls):
    with ThreadPoolExecutor(max_workers=5) as executor:  # Worker sayısını düşürdük
        results = list(executor.map(check_url, urls))
    return results

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/check_urls', methods=['POST'])
def check_urls():
    if 'file' in request.files:
        file = request.files['file']
        if file.filename.endswith('.xlsx'):
            df = pd.read_excel(file)
            urls = []
            if 'A' in df.columns:
                urls.extend(df['A'].dropna().tolist())
            if 'C' in df.columns:
                urls.extend(df['C'].dropna().tolist())
    else:
        urls = request.form.get('urls', '').split('\n')
        urls = [url.strip() for url in urls if url.strip()]

    results = process_urls(urls)
    return jsonify(results)

@app.route('/read_excel', methods=['POST'])
def read_excel():
    if 'file' not in request.files:
        return jsonify({'error': 'Dosya yüklenmedi'}), 400
    
    file = request.files['file']
    if not file.filename.endswith('.xlsx'):
        return jsonify({'error': 'Sadece Excel (.xlsx) dosyaları desteklenir'}), 400
    
    try:
        df = pd.read_excel(file)
        result = {
            'columns': [str(col) for col in df.columns],
            'data': df.fillna('').values.tolist()
        }
        return jsonify(result)
    except Exception as e:
        return jsonify({'error': f'Excel okuma hatası: {str(e)}'}), 500

@app.route('/export_excel', methods=['POST'])
def export_excel():
    try:
        data = request.json
        wb = openpyxl.Workbook()
        ws = wb.active
        
        # Başlıkları ekle
        for col, header in enumerate(data['headers'], start=1):
            ws.cell(row=1, column=col, value=header)
        
        # Verileri ekle ve renkleri ayarla
        green_fill = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')
        red_fill = PatternFill(start_color='FFB6C1', end_color='FFB6C1', fill_type='solid')
        
        for row_idx, row_data in enumerate(data['rows'], start=2):
            for col_idx, cell_data in enumerate(row_data, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=cell_data.get('text', ''))
                if cell_data.get('working') is True:
                    cell.fill = green_fill
                elif cell_data.get('working') is False:
                    cell.fill = red_fill
        
        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        
        return send_file(
            excel_file,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name='link_kontrol_sonuclari.xlsx'
        )
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# Development sunucusu için
if __name__ == '__main__':
    app.run(debug=False)  # Production'da debug kapalı 
